import openpyxl
import os
import pandas as pd
import utils.utils as u
from openpyxl.styles import Alignment, Font

from classes.environment_variable import EnvironmentVariable
from classes.menu import Menu


class MenuCollection(object):
    def __init__(self):
        u.clear_screen()
        self.get_folders()
        self.shopping_folder = str(EnvironmentVariable("shopping_folder", "string", False).value)
        self.menus = []

    def get_folders(self):
        self.temp_folder = u.make_or_get_directory("resources", "temp")
        self.output_folder = u.make_or_get_directory("resources", "output")

    def get_weekly_menus(self):
        self.menu_files = u.find_docx_files(self.shopping_folder)

    def analyse_menu(self):
        for menu_file in self.menu_files:
            menu = Menu(menu_file)
            menu.analyse()
            self.menus.append(menu)

    def extract_and_sort_recipes(self):
        all_recipes = []
        for menu in self.menus:
            for recipe in menu.recipes:
                all_recipes.append(recipe)

        # Sort recipes by recipe_date
        sorted_recipes = sorted(all_recipes, key=lambda r: r.recipe_date)

        # Prepare data for Excel (recipe_date, recipe, source)
        self.recipe_data = [{"recipe_date": recipe.recipe_date, "recipe": recipe.recipe, "source": recipe.source} for recipe in sorted_recipes]

    def write_to_excel(self):
        temp_filename = os.path.join(self.temp_folder, "temp.xlsx")
        output_filename = os.path.join(self.output_folder, "recipes.xlsx")

        # Convert the data to a DataFrame
        df = pd.DataFrame(self.recipe_data)

        # Write to Excel
        df.to_excel(temp_filename, index=False)

        # Load the Excel file where your recipes are stored
        df = pd.read_excel(temp_filename)

        # Create the pivot table for unique recipes and their count
        pivot_recipes = df.pivot_table(index="recipe", aggfunc=len, fill_value=0)

        # Create the pivot table for unique sources and their count
        pivot_sources = df.pivot_table(index="source", aggfunc=len, fill_value=0)

        # Reset index to make sure sorting works correctly
        pivot_recipes = pivot_recipes.reset_index()
        pivot_sources = pivot_sources.reset_index()

        # Sort both pivot tables by the count in descending order
        pivot_recipes = pivot_recipes.sort_values(by=[str(0), "recipe"], ascending=[False, True])
        pivot_sources = pivot_sources.sort_values(by=str(0), ascending=False)

        # Rename the column for counts
        pivot_recipes.columns = ["recipe", "Count"]
        pivot_sources.columns = ["source", "Count"]

        # Write both pivot tables to a new Excel file, each in a different sheet
        with pd.ExcelWriter(output_filename, engine="openpyxl") as writer:
            # Write the original data
            df.to_excel(writer, sheet_name="Original Data", index=False)

            # Write the first pivot table
            pivot_recipes.to_excel(writer, sheet_name="Recipe Count", index=False)

            # Write the second pivot table
            pivot_sources.to_excel(writer, sheet_name="Source Count", index=False)

        # Open the workbook to modify formatting
        wb = openpyxl.load_workbook(output_filename)

        # Function to apply formatting to a pivot table sheet
        def format_pivot_sheet(sheet_name):
            sheet = wb[sheet_name]

            # Set the width of the first column to 75
            sheet.column_dimensions["A"].width = 75

            # Left-align all content (starting from the second row, second column)
            for row in sheet.iter_rows(min_row=2, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal="left")
                    cell.font = Font(bold=False)  # Remove bold from all content

            # Bold the headers (first row)
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")  # Center-align header

        # Apply formatting to both pivot sheets
        format_pivot_sheet("Recipe Count")
        format_pivot_sheet("Source Count")

        # Save the modified workbook
        wb.save(output_filename)

        print("Pivot tables saved successfully.")
        os.remove(temp_filename)
